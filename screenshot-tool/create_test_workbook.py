"""テスト用Excelファイル生成スクリプト（openpyxl版）

スクリーンショットツールのPoC検証用に、以下の特徴を持つ.xlsxを生成:
- 複数シート（EstimateData, DeptSummary, Settings）
- 数式セル（SUM, COUNTIF, SUMIF, IF, ROUND）
- 書式のみ空セル（背景色あり・値なし → スマート領域検出テスト）
- 条件付き書式
- 名前付き範囲
- 固定行（Freeze Panes）

※ openpyxlではVBAマクロを含む.xlsm生成は限定的。
  VBAテスト用には別途既存.xlsmを使用するか、PS版を社内PCで実行すること。
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(OUTPUT_DIR, "TestWorkbook_HR_Estimate.xlsx")


def create_test_workbook():
    wb = Workbook()

    # --- 色定義 ---
    blue_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    gray_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    red_fill = PatternFill(start_color="FC9090", end_color="FC9090", fill_type="solid")
    bold_font = Font(bold=True)
    red_font = Font(color="990000")
    thin_border = Border(
        bottom=Side(style="thin")
    )

    # ================================================================
    # Sheet1: EstimateData（見積もりデータ）
    # ================================================================
    ws1 = wb.active
    ws1.title = "EstimateData"

    # ヘッダー行（行1）
    headers = ["No", "Name", "Dept", "Grade", "BaseSalary",
               "RoleAllow", "Commute", "Total", "MoM", "Note"]
    for c, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = bold_font
        cell.fill = blue_fill
        cell.border = thin_border

    # サブヘッダー（行2）
    sub_headers = ["", "", "", "", "(JPY)", "(JPY)", "(JPY)", "(JPY)", "(%)", ""]
    for c, sh in enumerate(sub_headers, 1):
        cell = ws1.cell(row=2, column=c, value=sh)
        cell.fill = green_fill

    # サンプルデータ（10行、行3～12）
    data = [
        (1, "Tanaka",    "Sales", "M1", 350000, 50000,  15000),
        (2, "Yamada",    "HR",    "S3", 280000, 0,      20000),
        (3, "Suzuki",    "Tech",  "E2", 320000, 30000,  12000),
        (4, "Sato",      "Sales", "M2", 380000, 60000,  18000),
        (5, "Takahashi", "Admin", "S2", 260000, 0,      10000),
        (6, "Watanabe",  "Tech",  "E3", 340000, 35000,  25000),
        (7, "Ito",       "HR",    "S1", 250000, 0,      8000),
        (8, "Nakamura",  "Sales", "M1", 350000, 50000,  15000),
        (9, "Kobayashi", "Tech",  "E1", 300000, 20000,  22000),
        (10,"Kato",      "Admin", "S3", 280000, 0,      10000),
    ]
    for i, (no, name, dept, grade, base, role, commute) in enumerate(data):
        row = i + 3
        ws1.cell(row=row, column=1, value=no)
        ws1.cell(row=row, column=2, value=name)
        ws1.cell(row=row, column=3, value=dept)
        ws1.cell(row=row, column=4, value=grade)
        ws1.cell(row=row, column=5, value=base).number_format = "#,##0"
        ws1.cell(row=row, column=6, value=role).number_format = "#,##0"
        ws1.cell(row=row, column=7, value=commute).number_format = "#,##0"
        # 合計: 数式
        ws1.cell(row=row, column=8).value = f"=E{row}+F{row}+G{row}"
        ws1.cell(row=row, column=8).number_format = "#,##0"
        # 前月比: 数式（ダミーで固定値から計算）
        ws1.cell(row=row, column=9).value = f"=ROUND((H{row}-H{row}*0.98)/H{row}*100,1)"
        ws1.cell(row=row, column=9).number_format = "0.0"

    # 合計行（行13）
    ws1.cell(row=13, column=4, value="Total").font = bold_font
    for c in range(5, 9):
        col_letter = get_column_letter(c)
        ws1.cell(row=13, column=c).value = f"=SUM({col_letter}3:{col_letter}12)"
        ws1.cell(row=13, column=c).number_format = "#,##0"
        ws1.cell(row=13, column=c).font = bold_font
        ws1.cell(row=13, column=c).fill = orange_fill

    # 書式のみの空セル（スマート領域検出テスト）
    # 行15-20、列A-J: 背景色のみ、値なし
    for r in range(15, 21):
        for c in range(1, 11):
            ws1.cell(row=r, column=c).fill = gray_fill

    # 条件付き書式: MoM > 2 → 赤背景
    ws1.conditional_formatting.add(
        "I3:I12",
        CellIsRule(operator="greaterThan", formula=["2"],
                   fill=red_fill, font=red_font)
    )

    # 固定行（行1-2を固定）
    ws1.freeze_panes = "A3"

    # 列幅調整
    col_widths = {"A": 5, "B": 14, "C": 8, "D": 8, "E": 12,
                  "F": 12, "G": 12, "H": 14, "I": 8, "J": 10}
    for col, width in col_widths.items():
        ws1.column_dimensions[col].width = width

    # ================================================================
    # Sheet2: DeptSummary（部署別サマリー）
    # ================================================================
    ws2 = wb.create_sheet("DeptSummary")

    dept_headers = ["Dept", "Count", "TotalCost", "AvgSalary"]
    for c, h in enumerate(dept_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = bold_font
        cell.fill = blue_fill

    depts = ["Sales", "HR", "Tech", "Admin"]
    for d, dept in enumerate(depts):
        row = d + 2
        ws2.cell(row=row, column=1, value=dept)
        # シート間参照数式
        ws2.cell(row=row, column=2).value = f"=COUNTIF(EstimateData!C:C,A{row})"
        ws2.cell(row=row, column=3).value = f"=SUMIF(EstimateData!C:C,A{row},EstimateData!H:H)"
        ws2.cell(row=row, column=3).number_format = "#,##0"
        ws2.cell(row=row, column=4).value = f"=IF(B{row}>0,C{row}/B{row},0)"
        ws2.cell(row=row, column=4).number_format = "#,##0"

    # 名前付き範囲
    dn1 = DefinedName("DeptMaster", attr_text="DeptSummary!$A$2:$A$5")
    wb.defined_names.add(dn1)
    dn2 = DefinedName("TotalCost", attr_text="DeptSummary!$C$2:$C$5")
    wb.defined_names.add(dn2)

    for col in ["A", "B", "C", "D"]:
        ws2.column_dimensions[col].width = 14

    # ================================================================
    # Sheet3: Settings（設定）
    # ================================================================
    ws3 = wb.create_sheet("Settings")

    ws3.cell(row=1, column=1, value="SettingName").font = bold_font
    ws3.cell(row=1, column=2, value="Value").font = bold_font

    settings = [
        ("CalcMonth", "2026-02"),
        ("TaxRate", 0.1),
        ("InsuranceRate", 0.15),
        ("CommuteLimit", 30000),
        ("Author", "BP_User"),
    ]
    for i, (name, val) in enumerate(settings):
        ws3.cell(row=i + 2, column=1, value=name)
        ws3.cell(row=i + 2, column=2, value=val)

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 14

    # ================================================================
    # 保存
    # ================================================================
    wb.save(OUTPUT_PATH)
    print(f"Test workbook created: {OUTPUT_PATH}")
    print(f"  Sheet1: EstimateData (freeze panes, formulas, conditional format, format-only empty cells)")
    print(f"  Sheet2: DeptSummary (cross-sheet formulas, named ranges)")
    print(f"  Sheet3: Settings (master data)")
    print(f"  Note: .xlsx format (no VBA). For VBA testing, use PS script on a PC with valid Excel license.")


if __name__ == "__main__":
    create_test_workbook()
